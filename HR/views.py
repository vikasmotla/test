from django.contrib.auth.models import User , Group
from django.shortcuts import render, redirect , get_object_or_404
from django.contrib.auth import authenticate , login , logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.urlresolvers import reverse
from django.template import RequestContext
from django.conf import settings as globalSettings
from django.core.exceptions import ObjectDoesNotExist , SuspiciousOperation
from django.views.decorators.csrf import csrf_exempt, csrf_protect
# Related to the REST Framework
from rest_framework import viewsets , permissions , serializers
from rest_framework.exceptions import *
from url_filter.integrations.drf import DjangoFilterBackend
from .serializers import *
from .models import *
from API.permissions import *
from ERP.models import application, permission
from ERP.views import getApps
from django.db.models import Q
from django.http import JsonResponse
import random, string
from django.utils import timezone
from rest_framework.views import APIView
from rest_framework.response import Response

from openpyxl import load_workbook
from io import BytesIO
from email.mime.application import MIMEApplication
from django.template.loader import render_to_string, get_template
from django.core.mail import send_mail, EmailMessage
import datetime



def sendOneLink(u, request):
    salt = hashlib.sha1(str(random.random())).hexdigest()[:5]
    key = hashlib.sha1(salt+u.email).hexdigest()
    key_expires = timezone.now() + datetime.timedelta(2)
    at = AuthToken(user = u , key = key , typ = 'linklogin' , valid = True , expiresOn = key_expires)
    at.save()
    link = globalSettings.SITE_ADDRESS + '/token/?key=' + key + '&next='+ request.data['next']

    msgBody = ['You are requested to fill your basic details on the link below.' , 'In case you have any query please contact us.']

    ctx = {
        'heading' : "Basic user details required",
        'recieverName' : u.first_name  + " " + u.last_name ,
        'message': msgBody,
        'link' : link,
        'linkUrl': 'vamso.cioc.co.in',
        'sendersAddress' : '(C) Vamso BioTec Pvt Ltd',
        'sendersPhone' : '122004',
        'linkedinUrl' : 'https://www.linkedin.com/',
        'fbUrl' : 'https://facebook.com',
        'twitterUrl' : 'https://twitter.com',
    }

    email_body = get_template('app.oneLink.email.html').render(ctx)

    msg = EmailMessage("[URGENT] Please update your account : " + u.first_name , email_body, to= [u.email, 'do_not_reply@cioc.co.in'], from_email= 'do_not_reply@cioc.co.in' )
    msg.content_subtype = 'html'
    msg.send()


class SendWelcomeEmail(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def post(self , request , format = None):
        u = User.objects.get(pk = request.data['user'])
        sendOneLink(u , request)
        return Response(status = status.HTTP_200_OK)

class ProfileStats(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def get(self , request , format = None):
        users =  []
        for u in User.objects.all():
            completion = 0
            # if u.profile.fathersName != None:
            #     completion += 10
            # if u.profile.localAddress != None:
            #     completion += 10
            # if u.profile.permanentAddress != None:
            #     completion += 10
            # if u.profile.prefix != 'NA':
            #     completion += 10
            # if u.profile.emergency != None:
            #     completion += 10

            users.append({ "mobile" : u.profile.mobile,"completion" : completion*2  , "name" : u.first_name + ' ' + u.last_name , "email" : u.email})

            print u.first_name , u.profile.mobile , u.email , u.username , completion*2 , '\n'

        ctx = {
            'heading' : "Profile Stats",
            'recieverName' : 'Admins' ,
            'users': users,
            'linkUrl': 'vamso.cioc.co.in',
            'sendersAddress' : '(C) Vamso BioTec Pvt Ltd',
            'sendersPhone' : '122004',
            'linkedinUrl' : 'https://www.linkedin.com/',
            'fbUrl' : 'https://facebook.com',
            'twitterUrl' : 'https://twitter.com',
        }

        email_body = get_template('app.profileStats.email.html').render(ctx)

        # print email_body

        msg = EmailMessage("Profile stats" , email_body, to= ['sarita.negi@vamsobiotec.com' , 'aakash.agarwal@vamsobiotec.com','do_not_reply@cioc.co.in', 'subhasis.ghosh@vamsobiotec.com'], from_email= 'do_not_reply@cioc.co.in' )
        msg.content_subtype = 'html'
        msg.send()

        return Response(status = status.HTTP_200_OK)



class SendOneLink(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def post(self , request , format = None):
        count = 0
        for u in User.objects.all():
            sendOneLink(u, request)
            count += 1
        return Response({"count" : count}, status = status.HTTP_200_OK)

class BulkUserAPIView(APIView):
    permission_classes = (permissions.IsAuthenticated , isAdmin)
    def post(self, request, format=None):
        print request.FILES
        wb = load_workbook(filename = BytesIO(request.FILES['xl'].read()))
        ws = wb.worksheets[0]
        row_count = ws.max_row
        column_count = ws.max_column

        count = 0

        for i in range(1, row_count):
            # sNo = ws['A' + str(i+1)].value
            first_name = ws['A' + str(i+1)].value
            last_name = ws['B' + str(i+1)].value
            email = ws['G' + str(i+1)].value
            dob = ws['E' + str(i+1)].value
            gender = ws['F' + str(i+1)].value
            mobile = ws['H' + str(i+1)].value

            if email is None:
                print "continue for row : " , i
                continue

            if dob is not None and dob.__class__ != unicode:
                dob = dob.strftime("%d%m%Y")

            count += 1;

            if 'vamsobiotec' in email:
                username = email.split('@')[0]
            else:
                print "generating the username"
                username = first_name.split(' ')[0].lower() + '.' + last_name.lower().replace(' ', '')[0]

            if User.objects.filter(username = username).count() >0:
                username += dob[-1]+ dob[0]

            print first_name , last_name , email , dob, gender, mobile, username , datetime.datetime.strptime(dob, "%d-%m-%Y")

            print '-------------\n'

            email = email.replace('\n' , '')

            user , created = User.objects.get_or_create(username = username)
            user.first_name = first_name
            user.last_name = last_name
            user.email = email
            user.set_password(dob.replace('-', ''))
            user.is_active = True
            user.save()


            p = user.profile
            p.dateOfBirth = datetime.datetime.strptime(dob, "%d-%m-%Y")
            p.gender = gender
            p.email = email
            p.mobile = mobile
            p.save()


            msgBody = ['Your account has been created, your username will be : <strong>%s</strong> <br> and password will be <strong>%s</strong>' %(user.username ,dob.replace('-', '') ), 'Upon login you are requested to change your password. You can find the option in the menu on the right side when you click your name.' , 'In case you have any query please contact us.']

            ctx = {
                'heading' : "New account creation",
                'recieverName' : first_name + " " + last_name,
                'message': msgBody,
                'linkUrl': 'cioc.co.in',
                'sendersAddress' : '(C) Vamso BioTec Pvt Ltd',
                'sendersPhone' : '122004',
                'linkedinUrl' : 'https://www.linkedin.com/',
                'fbUrl' : 'https://facebook.com',
                'twitterUrl' : 'https://twitter.com',
            }

            email_body = get_template('app.newAccount.email.html').render(ctx)
            msg = EmailMessage("New account creation : " + first_name , email_body, to= [email , 'do_not_reply@cioc.co.in'], from_email= 'do_not_reply@cioc.co.in' )
            msg.content_subtype = 'html'
            msg.send()

            add_application_access(user , ['app.dashboard' , 'app.calendar', 'app.myWork.MOM' , 'app.myWork'] , request.user)

        return Response({"count" : count}, status = status.HTTP_200_OK)

def generateOTPCode():
    length = 4
    chars = string.digits
    rnd = random.SystemRandom()
    return ''.join(rnd.choice(chars) for i in range(length))

def tokenAuthentication(request):
    ak = get_object_or_404(AuthToken, key=request.GET['key'] , typ='linklogin')
    #check if the activation key has expired, if it hase then render confirm_expired.html
    # if ak.expiresOn < timezone.now():
    #     raise SuspiciousOperation('Expired')
    #if the key hasn't expired save user and set him as active and render some template to confirm activation
    user = ak.user
    login(request , user,backend='django.contrib.auth.backends.ModelBackend')
    return redirect(request.GET['next'])

def generateOTP(request):
    print request.POST
    key_expires = timezone.now() + datetime.timedelta(2)
    otp = generateOTPCode()
    user = get_object_or_404(User, username = request.POST['id'])
    ak = accountsKey(user= user, activation_key= otp,
        key_expires=key_expires , keyType = 'otp')
    ak.save()
    print ak.activation_key
    # send a SMS with the OTP
    return JsonResponse({} ,status =200 )

def loginView(request):
    if globalSettings.LOGIN_URL != 'login':
        return redirect(reverse(globalSettings.LOGIN_URL))
    authStatus = {'status' : 'default' , 'message' : '' }
    statusCode = 200
    if request.user.is_authenticated():
        if request.GET:
            return redirect(request.GET['next'])
        else:
            return redirect(reverse(globalSettings.LOGIN_REDIRECT))
    if request.method == 'POST':

    	usernameOrEmail = request.POST['username']
        otpMode = False
        if 'otp' in request.POST:
            print "otp"
            otp = request.POST['otp']
            otpMode = True
        else:
            password = request.POST['password']
        if '@' in usernameOrEmail and '.' in usernameOrEmail:
            u = User.objects.get(email = usernameOrEmail)
            username = u.username
        else:
            username = usernameOrEmail
            try:
                u = User.objects.get(username = username)
            except:
                statusCode = 404
        if not otpMode:
            user = authenticate(username = username , password = password)
        else:
            print "OTP Mode"
            ak = None
            try:
                aks = accountsKey.objects.filter(activation_key=otp , keyType='otp')
                ak = aks[len(aks)-1]
                print "Aks", aks,ak
            except:
                pass
            print ak
            if ak is not None:
                #check if the activation key has expired, if it has then render confirm_expired.html
                if ak.key_expires > timezone.now():
                    user = ak.user
                    user.backend = 'django.contrib.auth.backends.ModelBackend'
                else:
                    user = None
            else:
                authStatus = {'status' : 'danger' , 'message' : 'Incorrect OTP'}
                statusCode = 401

    	if user is not None:
            login(request , user)
            if request.GET:
                return redirect(request.GET['next'])
            else:
                return redirect(reverse(globalSettings.LOGIN_REDIRECT))
        else:
            if statusCode == 200 and not u.is_active:
                authStatus = {'status' : 'warning' , 'message' : 'Your account is not active.'}
                statusCode = 423
            else:
                authStatus = {'status' : 'danger' , 'message' : 'Incorrect username or password.'}
                statusCode = 401

    return render(request , globalSettings.LOGIN_TEMPLATE , {'authStatus' : authStatus ,'useCDN' : globalSettings.USE_CDN , 'backgroundImage': globalSettings.LOGIN_PAGE_IMAGE}, status=statusCode)

def registerView(request):
    if globalSettings.REGISTER_URL != 'register':
        return redirect(reverse(globalSettings.REGISTER_URL))
    msg = {'status' : 'default' , 'message' : '' }
    if request.method == 'POST':
    	name = request.POST['name']
    	email = request.POST['email']
    	password = request.POST['password']
        if User.objects.filter(email = email).exists():
            msg = {'status' : 'danger' , 'message' : 'Email ID already exists' }
        else:
            user = User.objects.create(username = email.replace('@' , '').replace('.' ,''))
            user.first_name = name
            user.email = email
            user.set_password(password)
            user.save()
            user = authenticate(username = email.replace('@' , '').replace('.' ,'') , password = password)
            login(request , user)
            if request.GET:
                return redirect(request.GET['next'])
            else:
                return redirect(globalSettings.LOGIN_REDIRECT)
    return render(request , 'register.simple.html' , {'msg' : msg})


def logoutView(request):
    logout(request)
    return redirect(globalSettings.LOGOUT_REDIRECT)

def root(request):
    return redirect(globalSettings.ROOT_APP)

@login_required(login_url = globalSettings.LOGIN_URL)
def home(request):
    u = request.user
    if u.is_superuser:
        apps = application.objects.all()
    else:
        apps = getApps(u)

    apps = apps.filter(~Q(name__startswith='configure.' )).filter(~Q(name='app.users')).filter(~Q(name__endswith='.public'))

    return render(request , 'ngBase.html' , {'wampServer' : globalSettings.WAMP_SERVER, 'appsWithJs' : apps.filter(haveJs=True) \
    ,'appsWithCss' : apps.filter(haveCss=True) , 'useCDN' : globalSettings.USE_CDN , 'BRAND_LOGO' : globalSettings.BRAND_LOGO \
    ,'BRAND_NAME' :  globalSettings.BRAND_NAME})

class userProfileViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = userProfileSerializer
    queryset = profile.objects.all()

class OfficeViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    serializer_class = OficeSerializer
    queryset = Office.objects.all()
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['parent' ,]

class userProfileAdminModeViewSet(viewsets.ModelViewSet):
    permission_classes = (isAdmin ,)
    serializer_class = userProfileAdminModeSerializer
    queryset = profile.objects.all()

class UserProfileSelfModeViewSet(viewsets.ModelViewSet):
    permission_classes = (isOwner ,)
    serializer_class = UserProfileSelfModeSerializer
    def get_queryset(self):
        return profile.objects.filter(user = self.request.user)

class UserProfileSearchModeViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated ,)
    serializer_class = UserProfileSelfModeSerializer
    queryset = profile.objects.all()


class userAdminViewSet(viewsets.ModelViewSet):
    permission_classes = (isAdmin ,)
    queryset = User.objects.all()
    serializer_class = userAdminSerializer

class UserViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated ,)
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['username']
    serializer_class = userSerializer
    def get_queryset(self):
        if 'mode' in self.request.GET:
            if self.request.GET['mode']=="mySelf":
                if self.request.user.is_authenticated:
                    return User.objects.filter(username = self.request.user.username)
                else:
                    raise PermissionDenied()
            else :
                return User.objects.all().order_by('-date_joined')
        else:
            return User.objects.all().order_by('-date_joined')

class UserSearchViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated ,)
    filter_backends = [DjangoFilterBackend]
    filter_fields = ['username']
    serializer_class = userSearchSerializer
    queryset = User.objects.all()
    def get_queryset(self):
        if 'unit' in self.request.GET:
            return User.objects.filter(designation__in = Unit.objects.get(pk = self.request.GET['unit']).designations.all())
        if 'mode' in self.request.GET:
            if self.request.GET['mode']=="mySelf":
                if self.request.user.is_authenticated:
                    return User.objects.filter(username = self.request.user.username)
                else:
                    raise PermissionDenied()
            else :
                return User.objects.all().order_by('-date_joined')
        else:
            return User.objects.all().order_by('-date_joined')

class GroupViewSet(viewsets.ModelViewSet):
    permission_classes = (permissions.IsAuthenticated,)
    queryset = Group.objects.all()
    serializer_class = groupSerializer
